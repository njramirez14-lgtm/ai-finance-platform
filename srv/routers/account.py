import asyncio
import io
import json
import os
import re
from datetime import datetime, timedelta
from decimal import Decimal

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from sqlalchemy import func
from sqlalchemy.orm import Session

from srv.api.deps import get_db
from srv.api.deps_auth import get_current_user
from srv.models.account import Account
from srv.models.card import Card
from srv.models.category import Category
from srv.models.entity import Entity
from srv.models.subscription import Subscription
from srv.models.transaction import Transaction, TransactionType
from srv.schemas.account import (
    AccountCreate,
    AccountOut,
    AccountUpdate,
    BalanceAdjustment,
    CardMini,
)

router = APIRouter(prefix="/accounts", tags=["accounts"])

MAX_UPLOAD_BYTES = 8 * 1024 * 1024  # 8 MB
GEMINI_API_KEY = os.getenv("GOOGLE_API_KEY")
CHUNK_LINES = 100  # ~1 month of typical statements per chunk
HEADER_RE = re.compile(
    r"(fecha|date|importe|amount|concepto|description|saldo|balance|operaci[oó]n|movimiento)",
    re.IGNORECASE,
)


def _read_file_as_text(content: bytes, filename: str, mime: str) -> str:
    """Convert upload (PDF / Excel / CSV / text) to plain line-oriented text."""
    name = (filename or "").lower()
    mime_l = (mime or "").lower()

    if name.endswith(".pdf") or "pdf" in mime_l:
        try:
            from pypdf import PdfReader
        except ImportError as exc:
            raise HTTPException(status_code=500, detail="pypdf no instalado") from exc
        try:
            reader = PdfReader(io.BytesIO(content))
            return "\n".join((page.extract_text() or "") for page in reader.pages)
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"PDF ilegible: {exc}")

    if name.endswith((".xlsx", ".xlsm")) or "spreadsheet" in mime_l:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise HTTPException(status_code=500, detail="openpyxl no instalado") from exc
        try:
            wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
            rows: list[str] = []
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in ws.iter_rows(values_only=True):
                    rows.append("\t".join("" if v is None else str(v) for v in row))
            return "\n".join(rows)
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"Excel ilegible: {exc}")

    # Default: treat as text / CSV / TSV
    return content.decode("utf-8", errors="ignore")


def _split_pdf_into_chunks(content: bytes, pages_per_chunk: int = 3) -> list[bytes]:
    """Split a PDF byte blob into multiple smaller PDF byte blobs of at most
    `pages_per_chunk` pages each. Used so Gemini Vision doesn't truncate the
    JSON output on long statements."""
    try:
        from pypdf import PdfReader, PdfWriter
    except ImportError:
        return [content]
    try:
        reader = PdfReader(io.BytesIO(content))
    except Exception:
        return [content]
    total = len(reader.pages)
    if total <= pages_per_chunk:
        return [content]
    out: list[bytes] = []
    for start in range(0, total, pages_per_chunk):
        writer = PdfWriter()
        for i in range(start, min(start + pages_per_chunk, total)):
            writer.add_page(reader.pages[i])
        buf = io.BytesIO()
        writer.write(buf)
        out.append(buf.getvalue())
    return out


def _chunk_text(text: str, max_lines: int = CHUNK_LINES) -> list[str]:
    """Split text into ≤ max_lines chunks, preserving any column header line."""
    lines = [ln for ln in text.split("\n") if ln.strip()]
    if not lines:
        return []
    if len(lines) <= max_lines:
        return ["\n".join(lines)]
    first_line = lines[0]
    has_header = bool(HEADER_RE.search(first_line)) and len(first_line) < 300
    body = lines[1:] if has_header else lines
    chunks: list[str] = []
    for i in range(0, len(body), max_lines):
        chunk_lines = ([first_line] if has_header else []) + body[i : i + max_lines]
        chunks.append("\n".join(chunk_lines))
    return chunks


def _normalize_parsed(parsed) -> tuple[list[dict], float | None]:
    """Accept either a bare array of rows (legacy) or an object
    {closing_balance, transactions}. Returns (rows, closing_balance)."""
    if isinstance(parsed, list):
        return parsed, None
    if isinstance(parsed, dict):
        rows = parsed.get("transactions")
        rows = rows if isinstance(rows, list) else []
        cb = parsed.get("closing_balance")
        try:
            cb = float(cb) if cb is not None else None
        except (TypeError, ValueError):
            cb = None
        return rows, cb
    return [], None


def _extract_one(model, prompt: str) -> tuple[list[dict], float | None]:
    """Call Gemini once (sync). Returns (rows, closing_balance) (or raises)."""
    response = model.generate_content(prompt)
    raw = (response.text or "").strip()
    cleaned = raw.replace("```json", "").replace("```", "").strip()
    if not cleaned:
        return [], None
    return _normalize_parsed(json.loads(cleaned))


async def _extract_one_async(model, prompt: str, sem: asyncio.Semaphore) -> tuple[list[dict], float | None, str | None]:
    """Async variant with a concurrency semaphore.
    Returns (rows, closing_balance, error_kind). error_kind in {None, 'rate_limit', 'other'}."""
    async with sem:
        try:
            rows, cb = await asyncio.to_thread(_extract_one, model, prompt)
            return rows, cb, None
        except Exception as exc:
            msg = str(exc)
            if "429" in msg or "quota" in msg.lower() or "rate" in msg.lower():
                return [], None, "rate_limit"
            return [], None, "other"


def _verify_entity(db: Session, entity_id: int | None, user_id: int) -> None:
    if entity_id is None:
        return
    exists = (
        db.query(Entity)
        .filter(Entity.id == entity_id, Entity.user_id == user_id)
        .first()
    )
    if not exists:
        raise HTTPException(status_code=400, detail="Entity does not belong to user")


def _account_balance(db: Session, account: Account) -> Decimal:
    initial = Decimal(str(account.initial_balance or 0))
    income = (
        db.query(func.coalesce(func.sum(Transaction.amount), 0))
        .filter(
            Transaction.account_id == account.id,
            Transaction.type == TransactionType.INCOME,
        )
        .scalar()
        or 0
    )
    expense = (
        db.query(func.coalesce(func.sum(Transaction.amount), 0))
        .filter(
            Transaction.account_id == account.id,
            Transaction.type == TransactionType.EXPENSE,
        )
        .scalar()
        or 0
    )
    # TRANSFER rows are intentionally NOT subtracted: the user already tracks
    # the same money in the destination account (e.g. PayPal merchant
    # EXPENSEs) and double-counting would push the global patrimonio down
    # by the transfer amount on every payment. Use "Ajustar saldo" on the
    # source account to reconcile with the bank's real balance.
    return initial + Decimal(str(income)) - Decimal(str(expense))


def _account_monthly_flow(db: Session, account_id: int) -> tuple[Decimal, Decimal]:
    cutoff = datetime.utcnow() - timedelta(days=30)
    rows = (
        db.query(Transaction.type, func.coalesce(func.sum(Transaction.amount), 0))
        .filter(
            Transaction.account_id == account_id,
            Transaction.date >= cutoff,
        )
        .group_by(Transaction.type)
        .all()
    )
    by_type = {t: Decimal(str(amt)) for t, amt in rows}
    income = by_type.get(TransactionType.INCOME, Decimal("0"))
    expense = by_type.get(TransactionType.EXPENSE, Decimal("0"))
    return income, expense


def _to_out(db: Session, account: Account) -> dict:
    cards = (
        db.query(Card)
        .filter(Card.account_id == account.id, Card.user_id == account.user_id)
        .all()
    )
    monthly_income, monthly_expense = _account_monthly_flow(db, account.id)
    return {
        "id": account.id,
        "user_id": account.user_id,
        "name": account.name,
        "type": account.type,
        "currency": account.currency,
        "initial_balance": Decimal(str(account.initial_balance or 0)),
        "entity_id": account.entity_id,
        "account_number": account.account_number,
        "notes": account.notes,
        "transfer_patterns": account.transfer_patterns,
        "created_at": account.created_at,
        "balance": _account_balance(db, account),
        "cards": [CardMini.model_validate(c) for c in cards],
        "monthly_income": monthly_income,
        "monthly_expense": monthly_expense,
    }


def _get_owned_account(db: Session, account_id: int, user_id: int) -> Account:
    account = (
        db.query(Account)
        .filter(Account.id == account_id, Account.user_id == user_id)
        .first()
    )
    if not account:
        raise HTTPException(status_code=404, detail="Account not found")
    return account


@router.post("/", response_model=AccountOut, status_code=status.HTTP_201_CREATED)
def create_account(
    payload: AccountCreate,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    _verify_entity(db, payload.entity_id, current_user.id)
    account = Account(
        user_id=current_user.id,
        entity_id=payload.entity_id,
        name=payload.name,
        type=payload.type,
        currency=payload.currency or "EUR",
        initial_balance=payload.initial_balance or Decimal("0"),
        account_number=payload.account_number,
        notes=payload.notes,
        transfer_patterns=payload.transfer_patterns,
    )
    db.add(account)
    db.commit()
    db.refresh(account)
    return _to_out(db, account)


@router.get("/", response_model=list[AccountOut])
def list_accounts(
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    accounts = (
        db.query(Account)
        .filter(Account.user_id == current_user.id)
        .order_by(Account.created_at.asc())
        .all()
    )
    return [_to_out(db, a) for a in accounts]


@router.get("/{account_id}", response_model=AccountOut)
def get_account(
    account_id: int,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    account = _get_owned_account(db, account_id, current_user.id)
    return _to_out(db, account)


@router.put("/{account_id}", response_model=AccountOut)
def update_account(
    account_id: int,
    payload: AccountUpdate,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    account = _get_owned_account(db, account_id, current_user.id)

    if payload.entity_id is not None:
        _verify_entity(db, payload.entity_id, current_user.id)
        account.entity_id = payload.entity_id
    for field in ("name", "type", "currency", "initial_balance", "account_number", "notes", "transfer_patterns"):
        v = getattr(payload, field, None)
        if v is not None:
            setattr(account, field, v)

    db.commit()
    db.refresh(account)
    return _to_out(db, account)


@router.delete("/{account_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_account(
    account_id: int,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    account = _get_owned_account(db, account_id, current_user.id)
    # Detach references so the FK does not block deletion. The transactions
    # and cards remain — just unlinked from this account, as the UI promises.
    db.query(Transaction).filter(
        Transaction.user_id == current_user.id,
        Transaction.account_id == account.id,
    ).update({Transaction.account_id: None}, synchronize_session=False)
    db.query(Card).filter(
        Card.user_id == current_user.id,
        Card.account_id == account.id,
    ).update({Card.account_id: None}, synchronize_session=False)
    db.query(Subscription).filter(
        Subscription.user_id == current_user.id,
        Subscription.account_id == account.id,
    ).update({Subscription.account_id: None}, synchronize_session=False)
    db.delete(account)
    db.commit()
    return None


@router.post("/{account_id}/adjust-balance", response_model=AccountOut)
def adjust_balance(
    account_id: int,
    payload: BalanceAdjustment,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """Create a correction transaction so the account balance matches the target."""
    account = _get_owned_account(db, account_id, current_user.id)
    current = _account_balance(db, account)
    delta = Decimal(str(payload.target_balance)) - current
    if delta == 0:
        return _to_out(db, account)

    tx_type = TransactionType.INCOME if delta > 0 else TransactionType.EXPENSE
    description = payload.description or "Ajuste de saldo"
    tx = Transaction(
        amount=float(abs(delta)),
        type=tx_type,
        description=description,
        date=datetime.utcnow(),
        user_id=current_user.id,
        account_id=account.id,
        entity_id=account.entity_id,
    )
    db.add(tx)
    db.commit()
    db.refresh(account)
    return _to_out(db, account)


PROMPT_TEMPLATE = """Eres un experto extractor y categorizador de transacciones bancarias en España.
Analiza el siguiente texto y devuelve UN JSON OBJECT (sin markdown) con esta estructura:
{{
  "closing_balance": 140.01,
  "transactions": [
    {{
      "amount": 100.50,
      "type": "EXPENSE",
      "description": "Mercadona compra",
      "date": "2026-05-09",
      "merchant": "Mercadona",
      "category": "Alimentación"
    }}
  ]
}}
Reglas:
- closing_balance: el SALDO DISPONIBLE / saldo actual de la CUENTA que muestre el extracto
  (etiquetas típicas: "Saldo disponible", "Saldo final", "Saldo a fecha", "Saldo actual").
  Es el saldo total de la cuenta, NO el saldo parcial que aparece junto a cada línea.
  Devuélvelo como número con signo. Si el extracto NO muestra un saldo disponible/actual de
  la cuenta, pon null.
- type: "EXPENSE" (cargo), "INCOME" (abono) o "TRANSFER". Usa TRANSFER para movimientos entre
  cuentas propias del mismo titular: "TRASPASO PROPIO", "TRASPASO ENTRE CUENTAS", o
  transferencias cuyo ordenante/beneficiario sea el propio titular de la cuenta. Las TRANSFER
  NO son ni ingreso ni gasto y no se categorizan.
- Valores negativos en el extracto → EXPENSE; positivos → INCOME (salvo que sea TRANSFER).
- amount siempre en positivo absoluto.
- date YYYY-MM-DD. Si falta, usa fecha de hoy.
- description: concepto bruto del banco, limpio (sin códigos basura).
- merchant: nombre comercial del establecimiento si es identificable (Mercadona, Repsol, Netflix, Glovo, etc.).
- category: una sola categoría en español. Usa esta taxonomía (elige la más cercana, NO inventes nuevas):
  - Alimentación (supermercados, fruterías, panaderías)
  - Restauración (cafeterías, bares, restaurantes, Glovo/UberEats)
  - Transporte (gasolina, parking, taxi, Uber, Cabify, metro, bus, peaje)
  - Vivienda (alquiler, hipoteca, comunidad, IBI)
  - Suministros (luz, agua, gas, internet, móvil)
  - Suscripciones (Netflix, HBO, Spotify, Amazon Prime, Disney+, iCloud, software SaaS)
  - Salud (farmacia, médico, dentista, gimnasio, óptica)
  - Ocio (cine, conciertos, viajes, hoteles, vacaciones)
  - Compras (ropa, Amazon, electrónica, hogar)
  - Educación (cursos, libros, formación)
  - Seguros (auto, hogar, salud, vida)
  - Comisiones (bancarias, transferencias)
  - Impuestos (IRPF, IVA, Hacienda, multas)
  - Transferencia (Bizum, transferencia recibida/enviada sin destino claro)
  - Nómina (salario, paga extra)
  - Inversión (compra/venta acciones, ETFs, cripto, fondos)
  - Otros (sólo si nada encaja)
- Si no encuentras transacciones devuelve {{"closing_balance": null, "transactions": []}}.

Texto:
{text}
"""


@router.post("/{account_id}/upload-statement")
async def upload_statement(
    account_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """Parse a bank statement (CSV, TXT, PDF, XLSX) chunk-by-chunk with Gemini
    and persist transactions linked to this account. Resilient: partial chunk
    failures still commit the rest."""
    if not GEMINI_API_KEY:
        raise HTTPException(status_code=500, detail="Gemini API Key not configured")

    account = _get_owned_account(db, account_id, current_user.id)
    content = await file.read()
    if len(content) > MAX_UPLOAD_BYTES:
        raise HTTPException(status_code=413, detail="Archivo demasiado grande (máx 8 MB)")

    text_content = _read_file_as_text(content, file.filename or "", file.content_type or "")
    name_l = (file.filename or "").lower()
    mime_l = (file.content_type or "").lower()
    is_pdf = name_l.endswith(".pdf") or "pdf" in mime_l

    import google.generativeai as genai

    genai.configure(api_key=GEMINI_API_KEY)
    model = genai.GenerativeModel(
        "gemini-2.5-flash",
        generation_config={
            "temperature": 0.1,
            "max_output_tokens": 32768,
            "response_mime_type": "application/json",
        },
    )

    # If text extraction failed (scanned / compressed PDF that stripped the
    # text layer), send the raw PDF to Gemini Vision page-by-page so each
    # chunk's JSON output stays well under the 32k-token cap.
    used_vision = False
    if is_pdf and len(text_content.strip()) < 100:
        pdf_chunks = _split_pdf_into_chunks(content, pages_per_chunk=3)

        async def _vision_one(pdf_bytes: bytes, sem: asyncio.Semaphore) -> tuple[list[dict], float | None, str | None]:
            def _call() -> tuple[list[dict], float | None]:
                resp = model.generate_content([
                    {"mime_type": "application/pdf", "data": pdf_bytes},
                    PROMPT_TEMPLATE.format(text="(see attached PDF)"),
                ])
                raw = (resp.text or "").strip().replace("```json", "").replace("```", "").strip()
                if not raw:
                    return [], None
                return _normalize_parsed(json.loads(raw))
            async with sem:
                try:
                    rows, cb = await asyncio.to_thread(_call)
                    return rows, cb, None
                except json.JSONDecodeError:
                    return [], None, "other"
                except Exception as exc:
                    msg = str(exc)
                    if "429" in msg or "quota" in msg.lower() or "rate" in msg.lower():
                        return [], None, "rate_limit"
                    return [], None, "other"

        sem = asyncio.Semaphore(10)
        tasks = [_vision_one(c, sem) for c in pdf_chunks]
        try:
            results = await asyncio.wait_for(asyncio.gather(*tasks), timeout=260.0)
        except asyncio.TimeoutError:
            results = [t.result() if t.done() and not t.exception() else ([], None, "other") for t in tasks]
        used_vision = True
        chunks = [f"(pdf-vision page {i*3+1})" for i in range(len(pdf_chunks))]
    else:
        if not text_content.strip():
            raise HTTPException(
                status_code=400,
                detail="No se ha podido extraer texto del archivo. Si es un PDF escaneado o muy comprimido, vuelve a exportarlo del banco sin re-comprimir.",
            )
        chunks = _chunk_text(text_content, max_lines=CHUNK_LINES)
        if not chunks:
            raise HTTPException(status_code=400, detail="No hay líneas reconocibles")

    if not used_vision:
        # Process chunks in parallel. With Gemini Tier 1 (1000 RPM) we can burst
        # 20 concurrent calls easily. maxDuration is 300s in vercel.json so even
        # very large statements (~25 chunks) fit comfortably.
        MAX_CONCURRENCY = 20
        sem = asyncio.Semaphore(MAX_CONCURRENCY)
        tasks = [
            _extract_one_async(model, PROMPT_TEMPLATE.format(text=ct), sem)
            for ct in chunks
        ]
        try:
            results = await asyncio.wait_for(asyncio.gather(*tasks), timeout=260.0)
        except asyncio.TimeoutError:
            # Something hung; collect whatever finished
            results = [t.result() if t.done() and not t.exception() else ([], None, "other") for t in tasks]

    parsed: list[dict] = []
    failed_chunks = 0
    rate_limited = False
    closing_balance: float | None = None
    for rows, cb, err in results:
        if err == "rate_limit":
            rate_limited = True
            continue
        if err == "other":
            failed_chunks += 1
            continue
        parsed.extend(rows)
        # Keep the first available-balance the model reported (the header/page
        # that shows "Saldo disponible"); used below to pin the real balance.
        if closing_balance is None and cb is not None:
            closing_balance = cb

    if not parsed and rate_limited:
        raise HTTPException(
            status_code=429,
            detail="Has agotado la cuota gratuita de Gemini. Espera unos minutos o usa una API key de pago.",
        )

    # Preload categories for matching, keyed by lowercase name + type
    existing_categories = (
        db.query(Category).filter(Category.user_id == current_user.id).all()
    )
    cat_index: dict[tuple[str, TransactionType], Category] = {
        (c.name.lower(), c.type): c for c in existing_categories
    }

    def _resolve_category(name: str | None, tx_type: TransactionType) -> int | None:
        if not name:
            return None
        clean = name.strip()
        if not clean:
            return None
        key = (clean.lower(), tx_type)
        cat = cat_index.get(key)
        if cat:
            return cat.id
        # Use a savepoint so a unique-constraint violation here doesn't
        # roll back previously inserted transactions in this batch.
        sp = db.begin_nested()
        cat = Category(name=clean, type=tx_type, user_id=current_user.id)
        db.add(cat)
        try:
            db.flush()
            sp.commit()
        except Exception:
            sp.rollback()
            existing = (
                db.query(Category)
                .filter(
                    Category.user_id == current_user.id,
                    func.lower(Category.name) == clean.lower(),
                )
                .first()
            )
            if existing:
                cat_index[(existing.name.lower(), existing.type)] = existing
                return existing.id
            return None
        cat_index[key] = cat
        return cat.id

    # Internal-transfer keywords: account-configured patterns plus sensible
    # defaults so own-account movements (traspaso propio) aren't counted as
    # income/expense even when no patterns were configured.
    DEFAULT_TRANSFER_KEYWORDS = ["traspaso propio", "traspaso entre cuentas"]
    transfer_keywords: list[str] = list(DEFAULT_TRANSFER_KEYWORDS)
    if account.transfer_patterns:
        transfer_keywords += [
            p.strip().lower()
            for p in re.split(r"[,\n]+", account.transfer_patterns)
            if p.strip()
        ]

    def _is_transfer(text_desc: str | None) -> bool:
        if not transfer_keywords or not text_desc:
            return False
        low = text_desc.lower()
        return any(kw in low for kw in transfer_keywords)

    # Dedup against what's already stored for this account so re-uploading the
    # same (or an overlapping) statement doesn't double-count.
    existing_keys = {
        (d.date() if d else None, round(float(a or 0), 2), t, (ds or "")[:200])
        for d, a, t, ds in db.query(
            Transaction.date, Transaction.amount, Transaction.type, Transaction.description
        ).filter(Transaction.account_id == account.id).all()
    }
    skipped_duplicates = 0

    created = 0
    transfers_flagged = 0
    by_category: dict[str, int] = {}
    by_month: dict[str, dict] = {}
    by_type = {"INCOME": 0, "EXPENSE": 0, "TRANSFER": 0}
    income_total = 0.0
    expense_total = 0.0
    transfer_total = 0.0
    for item in parsed:
        if not isinstance(item, dict):
            continue
        try:
            amount = float(item.get("amount", 0))
            tx_type_raw = str(item.get("type", "EXPENSE")).upper()
            if tx_type_raw == "TRANSFER":
                tx_type = TransactionType.TRANSFER
            elif tx_type_raw == "INCOME":
                tx_type = TransactionType.INCOME
            else:
                tx_type = TransactionType.EXPENSE
            raw_desc = str(item.get("description") or "").strip()
            merchant = str(item.get("merchant") or "").strip()
            if merchant and raw_desc and merchant.lower() not in raw_desc.lower():
                desc = f"{merchant} — {raw_desc}"[:200]
            else:
                desc = (merchant or raw_desc)[:200] or None
            date_str = item.get("date")
            tx_date = datetime.utcnow()
            if date_str:
                try:
                    tx_date = datetime.fromisoformat(str(date_str))
                except ValueError:
                    pass
            cat_name = str(item.get("category") or "").strip() or None
        except (TypeError, ValueError):
            continue
        if amount <= 0:
            continue
        # If the line matches a transfer pattern (own-account movement), demote
        # INCOME/EXPENSE → TRANSFER so it doesn't count as real income/expense
        # (it's the same money moving between the user's own accounts).
        if tx_type != TransactionType.TRANSFER and _is_transfer(desc):
            tx_type = TransactionType.TRANSFER
        if tx_type == TransactionType.TRANSFER:
            cat_name = None  # transfers aren't categorized
        # Skip duplicates (same day, amount, type, description).
        dup_key = (tx_date.date(), round(amount, 2), tx_type, (desc or "")[:200])
        if dup_key in existing_keys:
            skipped_duplicates += 1
            continue
        existing_keys.add(dup_key)
        if tx_type == TransactionType.TRANSFER:
            transfers_flagged += 1
        category_id = _resolve_category(cat_name, tx_type) if tx_type != TransactionType.TRANSFER else None
        db.add(Transaction(
            amount=amount,
            type=tx_type,
            description=desc,
            date=tx_date,
            user_id=current_user.id,
            account_id=account.id,
            entity_id=account.entity_id,
            category_id=category_id,
        ))
        created += 1
        if cat_name and tx_type != TransactionType.TRANSFER:
            by_category[cat_name] = by_category.get(cat_name, 0) + 1
        if tx_type == TransactionType.INCOME:
            by_type["INCOME"] += 1
            income_total += amount
        elif tx_type == TransactionType.TRANSFER:
            by_type["TRANSFER"] += 1
            transfer_total += amount
        else:
            by_type["EXPENSE"] += 1
            expense_total += amount
        month_key = tx_date.strftime("%Y-%m")
        bucket = by_month.setdefault(
            month_key,
            {"count": 0, "income": 0.0, "expense": 0.0, "transfer": 0.0},
        )
        bucket["count"] += 1
        if tx_type == TransactionType.INCOME:
            bucket["income"] += amount
        elif tx_type == TransactionType.TRANSFER:
            bucket["transfer"] += amount
        else:
            bucket["expense"] += amount
    db.commit()
    db.refresh(account)

    # Pin the account balance to the statement's real available balance, so the
    # total comes out right automatically — independent of which period was
    # imported, duplicates, or transfer mis-tagging. Since
    # balance = initial_balance + income - expense (transfers excluded), we set
    # initial_balance = closing_balance - (income - expense).
    balance_pinned = None
    if closing_balance is not None:
        inc = db.query(func.coalesce(func.sum(Transaction.amount), 0)).filter(
            Transaction.account_id == account.id,
            Transaction.type == TransactionType.INCOME,
        ).scalar() or 0
        exp = db.query(func.coalesce(func.sum(Transaction.amount), 0)).filter(
            Transaction.account_id == account.id,
            Transaction.type == TransactionType.EXPENSE,
        ).scalar() or 0
        net = Decimal(str(inc)) - Decimal(str(exp))
        account.initial_balance = Decimal(str(closing_balance)) - net
        db.commit()
        db.refresh(account)
        balance_pinned = float(closing_balance)

    rate_limited_count = sum(1 for *_, e in results if e == "rate_limit")
    return {
        "success": True,
        "imported": created,
        "skipped_duplicates": skipped_duplicates,
        "balance_pinned": balance_pinned,
        "transfers_flagged": transfers_flagged,
        "used_vision": used_vision,
        "chunks_processed": len(chunks) - failed_chunks - rate_limited_count,
        "chunks_total": len(chunks),
        "failed_chunks": failed_chunks,
        "rate_limited": rate_limited,
        "by_category": by_category,
        "by_type": by_type,
        "by_month": dict(sorted(by_month.items(), reverse=True)),
        "income_total": round(income_total, 2),
        "expense_total": round(expense_total, 2),
        "transfer_total": round(transfer_total, 2),
        "account": _to_out(db, account),
    }


@router.post("/reconcile-transfers")
def reconcile_transfers(
    days_window: int = 3,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """For every unlinked TRANSFER row owned by this user, look for a matching
    EXPENSE in a different account with the same amount within ±days_window
    days. When exactly one candidate exists, link them via
    transactions.linked_transaction_id."""
    transfers = (
        db.query(Transaction)
        .filter(
            Transaction.user_id == current_user.id,
            Transaction.type == TransactionType.TRANSFER,
            Transaction.linked_transaction_id.is_(None),
        )
        .all()
    )
    linked = 0
    ambiguous = 0
    no_match = 0
    for t in transfers:
        if not t.date or t.amount is None:
            continue
        lo = t.date - timedelta(days=days_window)
        hi = t.date + timedelta(days=days_window)
        # Amount must match within 1 cent.
        candidates = (
            db.query(Transaction)
            .filter(
                Transaction.user_id == current_user.id,
                Transaction.type == TransactionType.EXPENSE,
                Transaction.account_id != t.account_id,
                Transaction.amount >= float(t.amount) - 0.01,
                Transaction.amount <= float(t.amount) + 0.01,
                Transaction.date >= lo,
                Transaction.date <= hi,
                Transaction.linked_transaction_id.is_(None),
            )
            .all()
        )
        if len(candidates) == 1:
            other = candidates[0]
            t.linked_transaction_id = other.id
            other.linked_transaction_id = t.id
            linked += 1
        elif len(candidates) > 1:
            ambiguous += 1
        else:
            no_match += 1
    db.commit()
    return {
        "linked_pairs": linked,
        "ambiguous": ambiguous,
        "no_match": no_match,
        "transfers_scanned": len(transfers),
    }
